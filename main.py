import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from datetime import datetime
import time
from selenium.webdriver.chrome.options import Options
import openpyxl

# Configurar o caminho do ChromeDriver
driver_path = r"C:\Users\Public\Chromedriver\chromedriver-win64\chromedriver.exe"  # Atualize para o caminho correto
service = Service(driver_path)

# URL do site de previsão do tempo
url = "https://www.climatempo.com.br/"

# Nome do arquivo Excel para salvar os dados
arquivo_excel = "dados_temperatura.xlsx"

# Função para criar a planilha Excel (caso não exista)
def criar_arquivo_excel():
    try:
        wb = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados Clima"
        ws.append(["Data e Hora", "Temperatura", "Umidade"])  # Cabeçalhos
        wb.save(arquivo_excel)

# Função para salvar dados na planilha Excel
def salvar_dados_excel(data_hora, temperatura, umidade):
    wb = openpyxl.load_workbook(arquivo_excel)
    ws = wb.active
    ws.append([data_hora, temperatura, umidade])  # Adicionar dados à planilha
    wb.save(arquivo_excel)

# Função para capturar os dados
def capturar_dados():
    driver = None
    try:
        # Criar a planilha Excel se for a primeira execução
        criar_arquivo_excel()

        # Configuração para ignorar erros SSL e desabilitar permissão de localização
        chrome_options = Options()
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--incognito")  # Usar modo incógnito para evitar caches
        chrome_options.add_argument("--disable-geolocation")  # Desabilitar a geolocalização
        chrome_options.add_argument("--disable-notifications")  # Desabilitar notificações

        # Configuração para ignorar a permissão de localização
        prefs = {"profile.default_content_setting_values.geolocation": 2}  # 2 = bloquear
        chrome_options.add_experimental_option("prefs", prefs)

        # Inicializar o navegador com as opções configuradas
        driver = webdriver.Chrome(service=service, options=chrome_options)

        # Acessar o site de previsão do tempo
        driver.get(url)
        time.sleep(5)  # Aguarde o carregamento do site

        # Capturar os elementos de temperatura e umidade
        temperatura_element = driver.find_element(By.ID, "current-weather-temperature")
        umidade_element = driver.find_element(By.XPATH, '//*[@id="current-weather-humidity"]')

        # Extrair os dados
        temperatura = temperatura_element.text
        umidade = umidade_element.text

        # Registrar data e hora
        data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Salvar no Excel
        salvar_dados_excel(data_hora, temperatura, umidade)

        # Atualizar os rótulos da interface gráfica com os dados
        label_dados.config(text=f"Temperatura: {temperatura}\nUmidade: {umidade}")
        label_data.config(text=f"Data e Hora: {data_hora}")

        # Exibir os dados no console
        print(f"Dados salvos: {data_hora}, {temperatura}, {umidade}")

    except Exception as e:
        # Exibir uma mensagem de erro caso algo dê errado
        print(f"Erro ao capturar dados: {str(e)}")
        messagebox.showerror("Erro", "Ocorreu um erro ao capturar os dados.")

    finally:
        if driver:
            # Fechar a guia do Google automaticamente após a execução
            driver.quit()

# Função para fechar a interface gráfica após captura dos dados
def fechar_interface():
    root.quit()

# Criar a interface gráfica
root = tk.Tk()
root.title("Captura de Dados de Temperatura e Umidade")

# Ajustar o tamanho da janela
root.geometry("400x300")

# Texto de instrução
label = tk.Label(root, text="Clique no botão abaixo para capturar os dados:")
label.pack(pady=10)

# Rótulos para mostrar os dados de temperatura, umidade e data/hora
label_dados = tk.Label(root, text="Temperatura e Umidade:", font=("Arial", 14))
label_dados.pack(pady=10)

label_data = tk.Label(root, text="Data e Hora:", font=("Arial", 12))
label_data.pack(pady=10)

# Botão para capturar dados
botao_capturar = tk.Button(root, text="Capturar Dados", command=capturar_dados, font=("Arial", 14))
botao_capturar.pack(pady=20)

# Iniciar a interface gráfica
root.protocol("WM_DELETE_WINDOW", fechar_interface)  # Garante que o programa se comporta corretamente ao fechar a janela
root.mainloop()
